
import os
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pdfplumber
from docx import Document
import csv as csv_module

# Export extraction functions for use in app.py
def extract_text_from_pdf(file_path):
    """Extract text from PDF file"""
    try:
        text = ""
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
        return text
    except Exception as e:
        raise Exception(f"Failed to parse PDF: {str(e)}")

def extract_text_from_docx(file_path):
    """Extract text from DOCX file"""
    try:
        doc = Document(file_path)
        text = "\n".join([paragraph.text for paragraph in doc.paragraphs])
        return text
    except Exception as e:
        raise Exception(f"Failed to parse DOCX: {str(e)}")

def extract_data_from_csv(file_path):
    """Extract data from CSV file"""
    try:
        data = []
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
            csv_reader = csv_module.DictReader(file)
            for row in csv_reader:
                data.append(dict(row))
        return data
    except Exception as e:
        raise Exception(f"Failed to parse CSV: {str(e)}")

def extract_text_from_txt(file_path):
    """Extract text from TXT file"""
    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
            return file.read()
    except Exception as e:
        raise Exception(f"Failed to read TXT file: {str(e)}")

def parse_text_data(text, filename):
    """Parse text data into structured format"""
    lines = [line.strip() for line in text.split('\n') if line.strip()]
    data = []
    
    # Try to extract structured data from text
    for i, line in enumerate(lines):
        if not line:
            continue
        
        # Try to parse key-value pairs
        colon_match = None
        equal_match = None
        
        if ':' in line:
            parts = line.split(':', 1)
            if len(parts) == 2:
                colon_match = (parts[0].strip(), parts[1].strip())
        
        if '=' in line and not colon_match:
            parts = line.split('=', 1)
            if len(parts) == 2:
                equal_match = (parts[0].strip(), parts[1].strip())
        
        if colon_match or equal_match:
            key, value = colon_match or equal_match
            data.append({
                'Property': key,
                'Value': value,
                'Source File': filename
            })
        elif line:
            # If no pattern matches, add as a row with text content
            data.append({
                'Content': line,
                'Source File': filename,
                'Line Number': i + 1
            })
    
    # If no structured data found, create a simple text extraction
    if not data and text.strip():
        words = text.strip().split()
        chunks = []
        chunk_size = 5
        for i in range(0, len(words), chunk_size):
            chunks.append(' '.join(words[i:i + chunk_size]))
        
        for idx, chunk in enumerate(chunks):
            data.append({
                'Extracted Text': chunk,
                'Source File': filename,
                'Segment': idx + 1
            })
    
    return data

def convert_files_to_excel(files, conversion_method='text-extraction', output_dir=None, output_filename=None):
    """Convert files to Excel format"""
    workbook = Workbook()
    workbook.remove(workbook.active)  # Remove default sheet
    
    # Resolve output directory
    resolved_output_dir = output_dir or 'output'
    os.makedirs(resolved_output_dir, exist_ok=True)
    
    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Process each file
    for file_info in files:
        file_path = file_info['path']
        original_filename = file_info['originalname']
        file_extension = Path(original_filename).suffix.lower()
        file_name_base = Path(original_filename).stem[:31]  # Excel sheet name limit
        
        print(f"Processing file: {original_filename} ({file_extension})")
        
        try:
            sheet_data = []
            
            if file_extension == '.pdf':
                text = extract_text_from_pdf(file_path)
                sheet_data = parse_text_data(text, original_filename)
            
            elif file_extension == '.docx':
                text = extract_text_from_docx(file_path)
                sheet_data = parse_text_data(text, original_filename)
            
            elif file_extension == '.csv':
                csv_data = extract_data_from_csv(file_path)
                sheet_data = [{
                    **row,
                    'Source File': original_filename
                } for row in csv_data]
            
            elif file_extension == '.txt':
                text = extract_text_from_txt(file_path)
                sheet_data = parse_text_data(text, original_filename)
            
            else:
                raise Exception(f"Unsupported file type: {file_extension}")
            
            # Create a worksheet for this file
            worksheet = workbook.create_sheet(title=file_name_base)
            
            # If we have data, add headers and rows
            if sheet_data:
                headers = list(sheet_data[0].keys())
                
                # Add header row
                header_row = worksheet.append(headers)
                for cell in header_row:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(vertical='center', horizontal='center')
                    cell.border = thin_border
                
                # Add data rows
                for row_data in sheet_data:
                    row_values = [row_data.get(header, '') for header in headers]
                    data_row = worksheet.append(row_values)
                    for cell in data_row:
                        cell.border = thin_border
                
                # Auto-fit columns
                for idx, header in enumerate(headers, 1):
                    column_letter = get_column_letter(idx)
                    max_length = max(
                        len(str(header)),
                        max([len(str(row_data.get(header, ''))) for row_data in sheet_data], default=0)
                    )
                    worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
            else:
                # If no data extracted, add a message
                worksheet.append(['No data extracted from file: ' + original_filename])
            
        except Exception as e:
            print(f"Error processing {original_filename}: {e}")
            # Create a worksheet with error message
            error_sheet_name = f"ERROR_{file_name_base[:26]}"
            error_sheet = workbook.create_sheet(title=error_sheet_name)
            error_sheet.append(['Error processing file: ' + original_filename])
            error_sheet.append(['Error message: ' + str(e)])
    
    # If no worksheets were created, create a default one
    if len(workbook.worksheets) == 0:
        workbook.create_sheet(title='No Data')
        workbook['No Data'].append(['No files were processed successfully'])
    
    # Generate or normalize output filename
    from datetime import datetime
    if output_filename:
        # Ensure it ends with .xlsx
        if not output_filename.lower().endswith('.xlsx'):
            output_filename = f"{output_filename}.xlsx"
    else:
        timestamp = datetime.now().strftime('%Y-%m-%dT%H-%M-%S')
        output_filename = f"converted_files_{timestamp}.xlsx"

    output_path = os.path.join(resolved_output_dir, output_filename)
    
    # Save the Excel file
    workbook.save(output_path)
    
    print(f"✅ Excel file created: {output_path}")
    
    return {
        'filename': output_filename,
        'filepath': output_path
    }

